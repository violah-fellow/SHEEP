from __future__ import annotations

import json
import os
from pydoc import text
import random
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zipfile import BadZipFile

import pandas as pd
import requests
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm


CONFIG: Dict[str, Any] = {
    "rules_text_file": "screening_rules.txt",
    "rules_yaml_file": "screening_rules.yaml",
    "workbook_file": "Copy of Test data for AI publication screening.xlsx",
    "output_file": "publication_screening_results.xlsx",
    "test_data_sheet": "Test data",
    "solutions_sheet": "Solutions",
    "id_column": "Publication ID",
    "title_column": "Title",
    "abstract_column": "Abstract",
    "backend": "gemini",  # "ollama", "anthropic", "gemini", or "openai"
    "ollama_base_url": "http://localhost:11434/v1",
    "ollama_api_key": "ollama",
    "ollama_model": "llama3.2",
    "anthropic_model": "claude-opus-4-6",
    "anthropic_api_key": None,
    "anthropic_max_tokens": 4096,
    "gemini_model": "gemini-3.1-pro-preview",
    "gemini_api_key": None,
    "openai_model": "gpt-4.1-mini",
    "openai_api_key": None,
    "openai_max_output_tokens": 4096,
    "batch_size": 10,
    "repetitions": 1,
    "temperature": 0.0,
    "sleep_between_calls_seconds": 2.0,
    "max_retries": 6,
    "request_timeout_seconds": 600,
    "retry_base_seconds": 5.0,
    "retry_max_seconds": 90.0,
    "checkpoint_dir": "benchmark_checkpoints",
    "resume_incomplete_run": True,
    "limit_rows": 300,
    "hide_technical_columns": True,
    "append_to_existing_workbook": True,
}

MAIN_DISPLAY_COLUMNS = [
    "Publication ID", "DOI", "PMID", "PMCID", "ISBN", "Title", "Abstract",
    "Scope", "Pillar", "Category", "Unsure", "Notes",
]

SUMMARY_VISIBLE_COLUMNS = [
    "run",
    "model",
    "n_records",
    "batch_size",
    "full_accuracy_normalized",
    "scope_accuracy",
    "pillar_accuracy_on_in_scope_normalized",
    "category_accuracy_on_in_scope_normalized",
    "wrong_scope",
    "right_scope_wrong_pillar",
    "right_scope_and_pillar_wrong_category",
    "unsure_rate",
]

GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")


@dataclass
class CallUsage:
    prompt_tokens: Optional[int] = None
    completion_tokens: Optional[int] = None
    total_tokens: Optional[int] = None
    estimated_cost_usd: Optional[float] = None


def load_env() -> None:
    load_dotenv()


def read_text_file(path: str) -> str:
    return Path(path).read_text(encoding="utf-8")


def read_yaml_file(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def normalize_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def find_column(df: pd.DataFrame, preferred: Optional[str], candidates: List[str]) -> Optional[str]:
    if preferred and preferred in df.columns:
        return preferred
    lowered = {str(c).strip().lower(): c for c in df.columns}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    return None


def normalize_scope(value: Any) -> str:
    val = normalize_text(value).lower()
    mappings = {
        "yes": "yes", "y": "yes", "in": "yes", "in scope": "yes", "true": "yes", "1": "yes",
        "no": "no", "n": "no", "out": "no", "out of scope": "no", "false": "no", "0": "no",
        "unsure": "unsure", "borderline": "unsure",
    }
    return mappings.get(val, val)


def normalize_solution_pillar(value: Any) -> str:
    raw = normalize_text(value)
    aliases = {
        "pb": "Plant-based", "plant-based": "Plant-based", "plant based": "Plant-based",
        "f": "Fermentation", "fermentation": "Fermentation",
        "cm": "Cultivated", "cultivated": "Cultivated", "cultivated meat": "Cultivated",
        "cc": "Cross-cutting", "cross-cutting": "Cross-cutting", "cross cutting": "Cross-cutting",
    }
    return aliases.get(raw.lower(), raw)


def normalize_solution_category(value: Any) -> str:
    raw = normalize_text(value)
    aliases = {
        "bioprocess design": "Bioprocess design",
        "cell line development": "Cell line development",
        "consumer and market research": "Consumer & market research",
        "crop development": "Crop development",
        "end product formulation": "End product formulation",
        "end product formulation & manufacturing": "End product formulation",
        "env/impact assessments": "Impact assessments",
        "feedstocks": "Feedstocks",
        "food safety & quality": "Food safety & quality",
        "health /nutrition": "Health & nutrition",
        "ingredient optimisation": "Ingredient optimisation",
        "manufacturing (incl texturization methods)": "Texturisation methods",
        "no technology sector assigned": "No category assigned",
        "scaffolding": "Scaffolding",
        "strain development": "Strain development",
        "target molecule selection": "Target molecule selection",
    }
    return aliases.get(raw.lower(), raw)


def normalize_predicted_pillar(value: Any) -> str:
    raw = normalize_text(value)
    aliases = {
        "plant-based": "Plant-based", "plant based": "Plant-based", "plant protein": "Plant-based", "pb": "Plant-based",
        "fermentation": "Fermentation", "precision fermentation": "Fermentation", "biomass fermentation": "Fermentation", "f": "Fermentation",
        "cultivated": "Cultivated", "cultivated meat": "Cultivated", "cultured meat": "Cultivated", "cell-cultivated": "Cultivated", "cm": "Cultivated",
    }
    return aliases.get(raw.lower(), raw)


def normalize_predicted_category(value: Any) -> str:
    raw = normalize_text(value)
    aliases = {
        "strain development": "Strain development",
        "cell line development": "Cell line development",
        "target molecule selection": "Target molecule selection",
        "cell culture media": "Cell culture media",
        "feedstocks": "Feedstocks",
        "bioprocess design": "Bioprocess design",
        "crop development": "Crop development",
        "ingredient optimisation": "Ingredient optimisation",
        "scaffolding": "Scaffolding",
        "texturisation methods": "Texturisation methods",
        "texturization methods": "Texturisation methods",
        "manufacturing": "Texturisation methods",
        "manufacturing (incl texturization methods)": "Texturisation methods",
        "end product formulation": "End product formulation",
        "end product formulation & manufacturing": "End product formulation",
        "health & nutrition": "Health & nutrition",
        "health and nutrition": "Health & nutrition",
        "health /nutrition": "Health & nutrition",
        "food safety & quality": "Food safety & quality",
        "food safety and quality": "Food safety & quality",
        "consumer & market research": "Consumer & market research",
        "consumer and market research": "Consumer & market research",
        "impact assessments": "Impact assessments",
        "env/impact assessments": "Impact assessments",
        "no category assigned": "No category assigned",
        "no technology sector assigned": "No category assigned",
    }
    return aliases.get(raw.lower(), raw)




def make_scope_schema() -> dict:
    return {
        "type": "object",
        "properties": {
            "results": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "record_id": {"type": "string"},
                        "in_scope": {"type": "string", "enum": ["yes", "no"]},
                        "review_flag": {"type": "string", "enum": ["none", "unsure", "borderline"]},
                        "brief_reason": {"type": "string"},
                    },
                    "required": ["record_id", "in_scope", "review_flag", "brief_reason"],
                    "additionalProperties": False,
                },
            }
        },
        "required": ["results"],
        "additionalProperties": False,
    }


def make_pillar_schema() -> dict:
    return {
        "type": "object",
        "properties": {
            "results": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "record_id": {"type": "string"},
                        "pillar": {"type": "string", "enum": ["Plant-based", "Fermentation", "Cultivated"]},
                        "review_flag": {"type": "string", "enum": ["none", "unsure", "borderline"]},
                        "brief_reason": {"type": "string"},
                    },
                    "required": ["record_id", "pillar", "review_flag", "brief_reason"],
                    "additionalProperties": False,
                },
            }
        },
        "required": ["results"],
        "additionalProperties": False,
    }


def make_category_schema(valid_categories: List[str]) -> dict:
    return {
        "type": "object",
        "properties": {
            "results": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "record_id": {"type": "string"},
                        "research_category": {"type": "string", "enum": valid_categories},
                        "review_flag": {"type": "string", "enum": ["none", "unsure", "borderline"]},
                        "brief_reason": {"type": "string"},
                    },
                    "required": ["record_id", "research_category", "review_flag", "brief_reason"],
                    "additionalProperties": False,
                },
            }
        },
        "required": ["results"],
        "additionalProperties": False,
    }


def combine_review_flags(*flags: str) -> str:
    lowered = {normalize_text(f).lower() for f in flags if normalize_text(f)}
    if "unsure" in lowered:
        return "unsure"
    if "borderline" in lowered:
        return "borderline"
    return "none"


def ensure_checkpoint_dir() -> Path:
    path = Path(CONFIG.get("checkpoint_dir", "benchmark_checkpoints"))
    path.mkdir(parents=True, exist_ok=True)
    return path


def get_run_checkpoint_path(run_idx: int) -> Path:
    return ensure_checkpoint_dir() / f"run_{run_idx}_checkpoint.json"


def load_run_checkpoint(run_idx: int) -> Optional[dict]:
    path = get_run_checkpoint_path(run_idx)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def save_run_checkpoint(
    run_idx: int,
    stage: str,
    next_batch_num: int,
    scope_results: List[Dict[str, Any]],
    pillar_results: List[Dict[str, Any]],
    category_results: List[Dict[str, Any]],
    usage_rows: List[Dict[str, Any]],
) -> None:
    path = get_run_checkpoint_path(run_idx)
    payload = {
        "run_idx": run_idx,
        "stage": stage,
        "next_batch_num": next_batch_num,
        "scope_results": scope_results,
        "pillar_results": pillar_results,
        "category_results": category_results,
        "usage_rows": usage_rows,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def delete_run_checkpoint(run_idx: int) -> None:
    path = get_run_checkpoint_path(run_idx)
    if path.exists():
        path.unlink()


def get_resume_run_number() -> Optional[int]:
    if not CONFIG.get("resume_incomplete_run", True):
        return None
    checkpoint_dir = ensure_checkpoint_dir()
    nums = []
    for path in checkpoint_dir.glob("run_*_checkpoint.json"):
        m = re.fullmatch(r"run_(\d+)_checkpoint\.json", path.name)
        if m:
            nums.append(int(m.group(1)))
    return min(nums) if nums else None


def is_retryable_error(exc: Exception) -> bool:
    text = str(exc).upper()
    retry_markers = [
        "503", "UNAVAILABLE", "RESOURCE_EXHAUSTED", "429",
        "TIMEOUT", "TIMED OUT", "READTIMEOUT", "CONNECTTIMEOUT",
    ]
    return any(marker in text for marker in retry_markers)


def retry_sleep_seconds(attempt: int) -> float:
    base = float(CONFIG.get("retry_base_seconds", 5.0))
    max_sleep = float(CONFIG.get("retry_max_seconds", 90.0))
    sleep_s = min(max_sleep, base * (2 ** attempt))
    jitter = random.uniform(0.0, min(3.0, sleep_s * 0.2))
    return sleep_s + jitter


def load_input_data() -> Tuple[pd.DataFrame, pd.DataFrame, str, dict]:
    test_df = pd.read_excel(CONFIG["workbook_file"], sheet_name=CONFIG["test_data_sheet"])
    solutions_df = pd.read_excel(CONFIG["workbook_file"], sheet_name=CONFIG["solutions_sheet"])
    test_df.columns = [str(c).strip() for c in test_df.columns]
    solutions_df.columns = [str(c).strip() for c in solutions_df.columns]
    rules_text = read_text_file(CONFIG["rules_text_file"])
    rules_yaml = read_yaml_file(CONFIG["rules_yaml_file"])
    if CONFIG.get("limit_rows"):
        test_df = test_df.head(int(CONFIG["limit_rows"])).copy()
        solutions_df = solutions_df.head(int(CONFIG["limit_rows"])).copy()
    return test_df, solutions_df, rules_text, rules_yaml


def prepare_records(test_df: pd.DataFrame) -> pd.DataFrame:
    df = test_df.copy()
    id_col = find_column(df, CONFIG.get("id_column"), ["Publication ID", "id", "ID", "record_id"])
    title_col = find_column(df, CONFIG.get("title_column"), ["Title", "title"])
    abstract_col = find_column(df, CONFIG.get("abstract_column"), ["Abstract", "abstract", "summary"])
    if id_col is None:
        df["__record_id"] = [f"row_{i+1}" for i in range(len(df))]
        id_col = "__record_id"
    if title_col is None:
        df["__title_fallback"] = ""
        title_col = "__title_fallback"
    if abstract_col is None:
        raise ValueError("Could not find abstract column.")
    df["_record_id"] = df[id_col].apply(normalize_text)
    df["_title"] = df[title_col].apply(normalize_text)
    df["_abstract"] = df[abstract_col].apply(normalize_text)
    return df


def build_scope_system_prompt(rules_text: str) -> str:
    return f"""
You are screening publication records for an alternative protein benchmark.

PASS 1: SCOPE ONLY
For each record:
1. Decide whether the publication is IN SCOPE or OUT OF SCOPE.
2. Assign review_flag: none, unsure, or borderline.
3. Give a short brief_reason.

Always provide your best answer even if uncertain.

A paper is IN SCOPE only if it is meaningfully about alternative proteins or ingredients for human food, including plant-based, microbial/fermentation-derived, or cultivated proteins/ingredients, and is relevant to product development, ingredient functionality, processing, bioprocessing, media/feedstocks, consumer research, regulation, techno-economics, or impacts.

Key IN-SCOPE cases:
- Characterisation or optimisation of plant, algal, microbial, or cultivated animal cells/tissues as protein or ingredient sources for human food
- Processing effects on protein functionality or quality for food use
- Crop, strain, cell line, feedstock, media, or bioprocess optimisation relevant to alternative proteins
- Functional comparison of alternative protein ingredients/products with conventional animal proteins
- Biochemical properties relevant to flavour, aroma, nutrition, allergenicity, or functionality
- Consumer, societal, policy, regulatory, techno-economic, or impact research specifically about alternative proteins
- Hybrid/blended alternative protein products where results help improve the alternative protein component

Key OUT-OF-SCOPE cases:
- Broad plant-based diets or vegan diets without relevance to alternative protein product development
- General crop/agronomy/breeding papers where protein or alternative protein use is not the real focus
- General food papers on bread, pasta, biscuits, snacks, etc. unless clearly about substituting animal proteins
- Nutraceutical, bioactive peptide, medical nutrition, or vulnerable-population use cases unless clearly tied to alternative protein substitution
- Pet food or animal feed only
- General fermentation, fungi, or food processing papers without clear alternative protein relevance
- General underutilised-food or protein-rich-food papers with no clear animal-protein substitution relevance
- Anything outside the inclusion criteria

Full detailed rules:
{rules_text}
""".strip()


def build_pillar_system_prompt(rules_text: str) -> str:
    return f"""
You are assigning the SINGLE BEST PILLAR to already in-scope alternative protein publications.

PASS 2: PILLAR ONLY
For each record:
1. Assign exactly one pillar:
   - Plant-based
   - Fermentation
   - Cultivated
2. Assign review_flag: none, unsure, or borderline.
3. Give a short brief_reason.

Always provide your best answer even if uncertain.

Pillar disambiguation rules:
- Plant-based: the main substance or ingredient comes directly from plants; also includes papers on plant protein ingredients and traditional fermentation used to modify plant ingredients
- Fermentation: the main production route uses microorganisms to produce biomass, protein, or specific functional ingredients; includes precision fermentation and biomass fermentation
- Cultivated: the main production route uses animal cells or tissues grown directly

Prefer Plant-based when:
- the paper is mainly about plant ingredients, plant proteins, plant functionality, crop protein sources, plant processing, or plant-based product development
- microbes are only used in a traditional/secondary way to modify plant ingredients

Prefer Fermentation when:
- the paper is mainly about microbial strains, microbial production hosts, fermentation feedstocks, precision fermentation, biomass fermentation, or fermentation bioprocesses
- the microorganisms themselves are the main production platform

Prefer Cultivated when:
- the paper is mainly about animal cells, cultivated meat/seafood, media, scaffolds, differentiation, proliferation, or cultivated production processes

Choose exactly one pillar.
Do not return blank.
Use unsure/borderline if needed, but still choose one pillar.

Full detailed rules:
{rules_text}
""".strip()


def build_category_system_prompt(rules_text: str) -> str:
    return """
You are assigning research categories to already in-scope alternative protein publications.

PASS 3: CATEGORY ONLY
1. Use the provided pillar.
2. Assign exactly one best research category from the allowed list.
3. Assign review_flag: none, unsure, or borderline.

Always provide your best answer even if uncertain.
Do NOT leave the category blank unless the paper is impossible to classify.
If several categories seem plausible, choose the closest single best category and use unsure/borderline if needed.

Allowed research categories and how to interpret them:

1. Strain development
- Screening or optimising microbial strains to improve production pathways or substrate modification

2. Cell line development
- Sourcing, optimising, banking, stabilising, or improving cultivated animal cell lines

3. Target molecule selection
- Identifying or validating target molecules/ingredients to be produced via precision fermentation

4. Cell culture media
- Novel media components, growth factors, amino acids, or cost-reduction strategies for cultivated systems

5. Feedstocks
- Alternative feedstocks or media/feed components for fermentation systems

6. Bioprocess design
- Upstream/downstream processing, reactor design, process control, scale-up, monitoring, separation, purification

7. Crop development
- Breeding or increased use of underutilised protein crops for higher protein yield/functionality

8. Ingredient optimisation
- Protein fractionation, functionalisation, improved ingredient functionality, solubility, emulsification, gelling, sensory enhancement, nutritional enhancement of ingredients

9. Scaffolding
- Biomaterials/scaffolds supporting cell adherence, growth, differentiation, or structure in cultivated meat

10. Texturisation methods
- Extrusion, electrospinning, 3D printing, enzymatic texturisation, or similar texture-forming processes

11. End product formulation
- Product formulation, fat integration, shelf life, stability, sensory testing, fortification, nutritional formulation, finished product design/testing

12. Health & nutrition
- Dietary impacts, bioavailability, health effects, nutrition outcomes, systematic reviews on health/nutrition of alternative proteins

13. Food safety & quality
- Safety, toxicology, assays, regulatory testing methods, contamination or quality/safety validation

14. Consumer & market research
- Consumer behaviour, naming/nomenclature, willingness to buy, retail or food environment, market research, branding

15. Impact assessments
- Life cycle assessment, techno-economic analysis, environmental/economic/social/geopolitical impacts, policy intervention impacts

16. No category assigned
- Broad review, political/ethical/philosophical discussion, or a paper clearly in scope but not well captured by the categories above

Category selection rules:
- Choose the category that best matches the paper's main contribution, not just keywords
- Prefer the central technical contribution over background context
- For plant protein functionality papers, Ingredient optimisation is often more appropriate than Health & nutrition
- For process-engineering papers, prefer Bioprocess design over Ingredient optimisation
- For consumer/policy/TEA/LCA papers, prefer Consumer & market research or Impact assessments as appropriate
- For broad reviews spanning many categories, use No category assigned

Pillar compatibility reminders:
- Cultivated-only categories: Cell line development, Cell culture media, Scaffolding
- Fermentation-heavy categories: Strain development, Target molecule selection, Feedstocks, Bioprocess design
- Cross-pillar categories: Texturisation methods, End product formulation, Health & nutrition, Food safety & quality, Consumer & market research, Impact assessments
- Plant-heavy categories: Crop development, Ingredient optimisation

Use the provided pillar.
Return exactly one category.
Use unsure/borderline if needed, but still choose a category.
""".strip()


def build_scope_user_prompt(batch_records: List[Dict[str, Any]]) -> str:
    return (
        "Output exactly one JSON object and nothing before or after it.\n"
        "Classify each record independently.\n"
        "Return valid JSON only.\n"
        "Do not omit any record.\n"
        "Preserve record_id exactly.\n\n"
        'Schema:\n{"results":[{"record_id":"...","in_scope":"yes|no","review_flag":"none|unsure|borderline","brief_reason":"short explanation"}]}\n\n'
        "Records:\n"
        + json.dumps(batch_records, ensure_ascii=False, indent=2)
    )


def build_pillar_user_prompt(batch_records: List[Dict[str, Any]]) -> str:
    return (
        "Output exactly one JSON object and nothing before or after it.\n"
        "Assign exactly one best-fit pillar to each record.\n"
        "Return valid JSON only.\n"
        "Do not omit any record.\n"
        "Preserve record_id exactly.\n\n"
        'Schema:\n{"results":[{"record_id":"...","pillar":"Plant-based|Fermentation|Cultivated","review_flag":"none|unsure|borderline","brief_reason":"short explanation"}]}\n\n'
        "Records:\n"
        + json.dumps(batch_records, ensure_ascii=False, indent=2)
    )


def build_category_user_prompt(batch_records: List[Dict[str, Any]]) -> str:
    return (
        "Output exactly one JSON object and nothing before or after it.\n"
        "Assign exactly one best-fit category to each record.\n"
        "Return valid JSON only.\n"
        "Do not omit any record.\n"
        "Preserve record_id exactly.\n\n"
        'Schema:\n{"results":[{"record_id":"...","research_category":"","review_flag":"none|unsure|borderline","brief_reason":"short explanation"}]}\n\n'
        "Records:\n"
        + json.dumps(batch_records, ensure_ascii=False, indent=2)
    )


def standardize_scope_item(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "record_id": normalize_text(item.get("record_id") or item.get("id")),
        "raw_in_scope": normalize_scope(item.get("in_scope")),
        "raw_review_flag_scope": normalize_text(item.get("review_flag")).lower() or "none",
        "raw_brief_reason_scope": normalize_text(item.get("brief_reason")),
    }


def standardize_pillar_item(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "record_id": normalize_text(item.get("record_id") or item.get("id")),
        "raw_pillar": normalize_text(item.get("pillar")),
        "raw_review_flag_pillar": normalize_text(item.get("review_flag")).lower() or "none",
        "raw_brief_reason_pillar": normalize_text(item.get("brief_reason")),
    }


def standardize_category_item(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "record_id": normalize_text(item.get("record_id") or item.get("id")),
        "raw_research_category": normalize_text(item.get("research_category")),
        "raw_review_flag_category": normalize_text(item.get("review_flag")).lower() or "none",
        "raw_brief_reason_category": normalize_text(item.get("brief_reason")),
    }


def strip_code_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z0-9_\-]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
    return text.strip()


def extract_json_object(text: str) -> Dict[str, Any]:
    cleaned = strip_code_fences(text).strip()

    try:
        parsed = json.loads(cleaned)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    decoder = json.JSONDecoder()

    try:
        parsed, _ = decoder.raw_decode(cleaned)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    for i, ch in enumerate(cleaned):
        if ch == "{":
            try:
                parsed, _ = decoder.raw_decode(cleaned[i:])
                if isinstance(parsed, dict):
                    return parsed
            except json.JSONDecodeError:
                continue

    raise ValueError("Could not parse JSON object from model output")


def estimate_cost(usage: CallUsage) -> Optional[float]:
    return 0.0 if CONFIG["backend"] == "ollama" else usage.estimated_cost_usd


def call_ollama_openai_compatible(system_prompt: str, user_prompt: str) -> Tuple[str, CallUsage]:
    url = CONFIG["ollama_base_url"].rstrip("/") + "/chat/completions"
    payload = {
        "model": CONFIG["ollama_model"],
        "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
        "temperature": CONFIG["temperature"],
        "response_format": {"type": "json_object"},
    }
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {CONFIG['ollama_api_key']}"}
    resp = requests.post(url, json=payload, headers=headers, timeout=CONFIG["request_timeout_seconds"])
    resp.raise_for_status()
    data = resp.json()
    raw_usage = data.get("usage", {}) or {}
    return data["choices"][0]["message"]["content"], CallUsage(
        prompt_tokens=raw_usage.get("prompt_tokens"),
        completion_tokens=raw_usage.get("completion_tokens"),
        total_tokens=raw_usage.get("total_tokens"),
        estimated_cost_usd=0.0,
    )


def call_anthropic(system_prompt: str, user_prompt: str) -> Tuple[str, CallUsage]:
    api_key = CONFIG.get("anthropic_api_key") or os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("Missing Anthropic API key. Set ANTHROPIC_API_KEY or CONFIG['anthropic_api_key'].")

    payload = {
        "model": CONFIG["anthropic_model"],
        "max_tokens": CONFIG.get("anthropic_max_tokens", 4096),
        "temperature": CONFIG["temperature"],
        "system": system_prompt,
        "messages": [{"role": "user", "content": user_prompt}],
    }

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json=payload,
        timeout=CONFIG["request_timeout_seconds"],
    )
    resp.raise_for_status()
    data = resp.json()

    text_blocks = [b.get("text", "") for b in data.get("content", []) if b.get("type") == "text"]
    text = "\n".join([t for t in text_blocks if t]).strip()
    if not text:
        raise RuntimeError("Anthropic returned no text content.")

    raw_usage = data.get("usage", {}) or {}
    usage = CallUsage(
        prompt_tokens=raw_usage.get("input_tokens"),
        completion_tokens=raw_usage.get("output_tokens"),
        total_tokens=(
            (raw_usage.get("input_tokens") or 0) + (raw_usage.get("output_tokens") or 0)
            if raw_usage.get("input_tokens") is not None or raw_usage.get("output_tokens") is not None
            else None
        ),
        estimated_cost_usd=None,
    )
    return text, usage


def call_gemini(system_prompt: str, user_prompt: str, response_schema: dict) -> Tuple[str, CallUsage]:
    api_key = CONFIG.get("gemini_api_key") or os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    if not api_key:
        raise RuntimeError("Missing Gemini API key. Set GEMINI_API_KEY/GOOGLE_API_KEY or CONFIG['gemini_api_key'].")

    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": f"{system_prompt}\n\n{user_prompt}"}],
            }
        ],
        "generationConfig": {
            "temperature": CONFIG["temperature"],
            "responseMimeType": "application/json",
            "responseJsonSchema": response_schema,
        },
    }

    url = f"https://generativelanguage.googleapis.com/v1beta/models/{CONFIG['gemini_model']}:generateContent?key={api_key}"
    resp = requests.post(url, json=payload, timeout=CONFIG["request_timeout_seconds"])
    resp.raise_for_status()
    data = resp.json()

    candidates = data.get("candidates", [])
    if not candidates:
        raise RuntimeError(f"Gemini returned no candidates: {data}")

    parts = candidates[0].get("content", {}).get("parts", [])
    text = "\n".join([p.get("text", "") for p in parts if p.get("text")]).strip()
    if not text:
        raise RuntimeError("Gemini returned no text content.")

    usage_meta = data.get("usageMetadata", {}) or {}
    usage = CallUsage(
        prompt_tokens=usage_meta.get("promptTokenCount"),
        completion_tokens=usage_meta.get("candidatesTokenCount"),
        total_tokens=usage_meta.get("totalTokenCount"),
        estimated_cost_usd=None,
    )
    return text, usage


def call_openai(system_prompt: str, user_prompt: str, response_schema: dict) -> Tuple[str, CallUsage]:
    api_key = CONFIG.get("openai_api_key") or os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Missing OpenAI API key. Set OPENAI_API_KEY or CONFIG['openai_api_key'].")

    payload = {
        "model": CONFIG["openai_model"],
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "max_tokens": CONFIG.get("openai_max_output_tokens", 4096),
        "temperature": CONFIG["temperature"],
        "response_format": {
            "type": "json_schema",
            "json_schema": {
                "name": "publication_screening_result",
                "strict": True,
                "schema": response_schema,
            }
        },
    }

    resp = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=CONFIG["request_timeout_seconds"],
    )
    resp.raise_for_status()
    data = resp.json()

    text = data["choices"][0]["message"]["content"]
    if not text:
        raise RuntimeError(f"OpenAI returned no text content: {data}")

    usage_meta = data.get("usage", {}) or {}
    usage = CallUsage(
        prompt_tokens=usage_meta.get("prompt_tokens"),
        completion_tokens=usage_meta.get("completion_tokens"),
        total_tokens=usage_meta.get("total_tokens"),
        estimated_cost_usd=None,
    )
    return text, usage


def call_model(system_prompt: str, user_prompt: str, response_schema: Optional[dict] = None) -> Tuple[str, CallUsage]:
    last_error: Optional[Exception] = None
    for attempt in range(CONFIG["max_retries"] + 1):
        try:
            if CONFIG["backend"] == "ollama":
                return call_ollama_openai_compatible(system_prompt, user_prompt)
            if CONFIG["backend"] == "anthropic":
                return call_anthropic(system_prompt, user_prompt)
            if CONFIG["backend"] == "gemini":
                if response_schema is None:
                    raise ValueError("Gemini backend requires response_schema")
                return call_gemini(system_prompt, user_prompt, response_schema)
            if CONFIG["backend"] == "openai":
                if response_schema is None:
                    raise ValueError("OpenAI backend requires response_schema")
                return call_openai(system_prompt, user_prompt, response_schema)
            raise ValueError("CONFIG['backend'] must be 'ollama', 'anthropic', 'gemini', or 'openai'")
        except Exception as exc:
            last_error = exc
            if attempt >= CONFIG["max_retries"]:
                break
            if is_retryable_error(exc):
                sleep_s = retry_sleep_seconds(attempt)
                print(f"Retryable model error: {exc}. Sleeping {sleep_s:.1f}s before retry {attempt + 1}/{CONFIG['max_retries']}.")
                time.sleep(sleep_s)
            else:
                time.sleep(1.0)
    raise RuntimeError(f"Model call failed after retries: {last_error}")


def chunk_records(records: List[Dict[str, Any]], batch_size: int) -> List[List[Dict[str, Any]]]:
    return [records[i:i + batch_size] for i in range(0, len(records), batch_size)]


def recover_missing_results_for_batch(
    batch: List[Dict[str, Any]],
    results: List[Dict[str, Any]],
    pass_name: str,
    batch_num: int,
    run_idx: int,
    system_prompt: str,
    build_user_prompt_fn,
    standardize_item_fn,
    usage_rows: List[Dict[str, Any]],
    response_schema: Optional[dict] = None,
) -> List[Dict[str, Any]]:
    expected_ids = [item["record_id"] for item in batch]
    returned_ids = {normalize_text(r.get("record_id")) for r in results if normalize_text(r.get("record_id"))}
    missing_ids = [rid for rid in expected_ids if rid not in returned_ids]
    recovered: List[Dict[str, Any]] = []

    for missing_id in missing_ids:
        single_record = [item for item in batch if item["record_id"] == missing_id]
        retry_content, retry_usage = call_model(system_prompt, build_user_prompt_fn(single_record), response_schema=response_schema)
        retry_parsed = extract_json_object(retry_content)
        retry_items = [standardize_item_fn(item) for item in retry_parsed.get("results", [])]
        retry_items = [item for item in retry_items if normalize_text(item.get("record_id")) == missing_id]
        recovered.extend(retry_items)
        usage_rows.append({
            "run": f"Run {run_idx}",
            "pass": f"{pass_name}_retry",
            "batch_number": batch_num,
            "backend": CONFIG["backend"],
            "model": CONFIG["ollama_model"] if CONFIG["backend"] == "ollama" else (CONFIG["anthropic_model"] if CONFIG["backend"] == "anthropic" else (CONFIG["gemini_model"] if CONFIG["backend"] == "gemini" else CONFIG["openai_model"])),
            "records_in_batch": 1,
            "prompt_tokens": retry_usage.prompt_tokens,
            "completion_tokens": retry_usage.completion_tokens,
            "total_tokens": retry_usage.total_tokens,
            "estimated_cost_usd": estimate_cost(retry_usage),
        })

    merged = results + recovered
    by_id: Dict[str, Dict[str, Any]] = {}
    for item in merged:
        rid = normalize_text(item.get("record_id"))
        if rid and rid not in by_id:
            by_id[rid] = item

    still_missing = [rid for rid in expected_ids if rid not in by_id]
    if still_missing:
        raise ValueError(f"{pass_name} batch {batch_num}: missing results after retry for {still_missing}")
    return [by_id[rid] for rid in expected_ids]


def validate_and_repair_batch_results(
    batch: List[Dict[str, Any]],
    results: List[Dict[str, Any]],
    pass_name: str,
    batch_num: int,
    run_idx: int,
    system_prompt: str,
    build_user_prompt_fn,
    standardize_item_fn,
    usage_rows: List[Dict[str, Any]],
    response_schema: Optional[dict] = None,
) -> List[Dict[str, Any]]:
    expected_ids = [item["record_id"] for item in batch]
    expected_set = set(expected_ids)

    cleaned_results: List[Dict[str, Any]] = []
    seen_ids = set()

    for item in results:
        rid = normalize_text(item.get("record_id"))
        if not rid or rid not in expected_set or rid in seen_ids:
            continue
        seen_ids.add(rid)
        cleaned_results.append(item)

    returned_ids = {item["record_id"] for item in cleaned_results}
    missing_ids = [rid for rid in expected_ids if rid not in returned_ids]

    if not missing_ids:
        by_id = {item["record_id"]: item for item in cleaned_results}
        return [by_id[rid] for rid in expected_ids]

    recovered = recover_missing_results_for_batch(
        batch=[item for item in batch if item["record_id"] in missing_ids],
        results=[],
        pass_name=pass_name,
        batch_num=batch_num,
        run_idx=run_idx,
        system_prompt=system_prompt,
        build_user_prompt_fn=build_user_prompt_fn,
        standardize_item_fn=standardize_item_fn,
        usage_rows=usage_rows,
        response_schema=response_schema,
    )

    merged = cleaned_results + recovered
    by_id: Dict[str, Dict[str, Any]] = {}
    for item in merged:
        rid = normalize_text(item.get("record_id"))
        if rid in expected_set and rid not in by_id:
            by_id[rid] = item

    still_missing = [rid for rid in expected_ids if rid not in by_id]
    if still_missing:
        raise ValueError(f"{pass_name} batch {batch_num}: still missing record_ids after repair: {still_missing}")

    return [by_id[rid] for rid in expected_ids]


def normalize_three_pass_result(scope_item: dict, pillar_item: Optional[dict], category_item: Optional[dict], rules_yaml: dict) -> dict:
    valid_in_scope = set(rules_yaml["valid_values"]["in_scope"])
    valid_review_flags = set(rules_yaml["valid_values"]["review_flag"])
    valid_pillars = set(rules_yaml["valid_values"]["pillars"])
    valid_categories = {x["name"] for x in rules_yaml["research_categories"]}
    category_to_pillars = {x["name"]: set(x["relevant_pillars"]) for x in rules_yaml["research_categories"]}

    raw_in_scope = normalize_scope(scope_item.get("raw_in_scope"))
    raw_pillar = normalize_text(pillar_item.get("raw_pillar")) if pillar_item else ""
    raw_category = normalize_text(category_item.get("raw_research_category")) if category_item else ""

    raw_scope_flag = normalize_text(scope_item.get("raw_review_flag_scope")).lower() or "none"
    raw_pillar_flag = normalize_text(pillar_item.get("raw_review_flag_pillar")).lower() if pillar_item else ""
    raw_cat_flag = normalize_text(category_item.get("raw_review_flag_category")).lower() if category_item else ""

    raw_scope_reason = normalize_text(scope_item.get("raw_brief_reason_scope"))
    raw_pillar_reason = normalize_text(pillar_item.get("raw_brief_reason_pillar")) if pillar_item else ""
    raw_cat_reason = normalize_text(category_item.get("raw_brief_reason_category")) if category_item else ""

    notes: List[str] = []
    validation_status = "valid"

    pred_in_scope = raw_in_scope if raw_in_scope in valid_in_scope else "no"
    if raw_in_scope not in valid_in_scope:
        notes.append("Invalid in_scope value normalized to no")
        validation_status = "normalized"

    pred_pillar = normalize_predicted_pillar(raw_pillar)
    pred_category = normalize_predicted_category(raw_category)

    final_flag = combine_review_flags(raw_scope_flag, raw_pillar_flag, raw_cat_flag)
    if final_flag not in valid_review_flags:
        final_flag = "unsure"
        notes.append("Invalid review_flag normalized to unsure")
        validation_status = "normalized"

    if pred_in_scope == "no":
        pred_pillar = ""
        pred_category = ""
    else:
        if not pred_pillar:
            notes.append("Missing pillar for in-scope record")
            validation_status = "normalized"
            final_flag = "unsure"
        elif pred_pillar not in valid_pillars:
            notes.append("Predicted pillar invalid")
            validation_status = "normalized"
            final_flag = "unsure"

        if not pred_category:
            notes.append("Missing category for in-scope record")
            validation_status = "normalized"
            final_flag = "unsure"
        elif pred_category not in valid_categories:
            notes.append("Predicted category invalid")
            validation_status = "normalized"
            final_flag = "unsure"
        elif pred_pillar in valid_pillars and pred_category in valid_categories:
            if pred_pillar not in category_to_pillars.get(pred_category, set()):
                notes.append("Category incompatible with pillar")
                validation_status = "normalized"
                final_flag = "unsure"

    combined_reason = " | ".join([x for x in [raw_scope_reason, raw_pillar_reason, raw_cat_reason] if x])

    return {
        "record_id": normalize_text(scope_item.get("record_id")),
        "pred_in_scope": pred_in_scope,
        "pred_pillar": pred_pillar,
        "pred_research_category": pred_category,
        "pred_review_flag": final_flag,
        "pred_brief_reason": combined_reason,
        "validation_status": validation_status,
        "validation_notes": " | ".join(notes),
    }


def run_one_experiment(prepared_df: pd.DataFrame, rules_text: str, rules_yaml: dict, run_idx: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    scope_system = build_scope_system_prompt(rules_text)
    pillar_system = build_pillar_system_prompt(rules_text)
    category_system = build_category_system_prompt(rules_text)

    scope_schema = make_scope_schema()
    pillar_schema = make_pillar_schema()
    category_schema = make_category_schema([x["name"] for x in rules_yaml["research_categories"]])

    input_records = [{"record_id": row["_record_id"], "title": row["_title"], "abstract": row["_abstract"]} for _, row in prepared_df.iterrows()]

    checkpoint = load_run_checkpoint(run_idx)
    scope_results: List[Dict[str, Any]] = checkpoint.get("scope_results", []) if checkpoint else []
    pillar_results: List[Dict[str, Any]] = checkpoint.get("pillar_results", []) if checkpoint else []
    category_results: List[Dict[str, Any]] = checkpoint.get("category_results", []) if checkpoint else []
    usage_rows: List[Dict[str, Any]] = checkpoint.get("usage_rows", []) if checkpoint else []
    resume_stage = checkpoint.get("stage", "scope") if checkpoint else "scope"
    resume_batch = int(checkpoint.get("next_batch_num", 1)) if checkpoint else 1

    scope_batches = chunk_records(input_records, CONFIG["batch_size"])
    if resume_stage == "scope":
        for batch_num, batch in enumerate(tqdm(scope_batches, desc=f"Run {run_idx} pass1_scope"), start=1):
            if batch_num < resume_batch:
                continue
            content, usage = call_model(scope_system, build_scope_user_prompt(batch), response_schema=(scope_schema if CONFIG["backend"] in {"gemini", "openai"} else None))
            parsed = extract_json_object(content)
            results = [standardize_scope_item(item) for item in parsed.get("results", [])]
            results = validate_and_repair_batch_results(
                batch=batch,
                results=results,
                pass_name="pass1_scope",
                batch_num=batch_num,
                run_idx=run_idx,
                system_prompt=scope_system,
                build_user_prompt_fn=build_scope_user_prompt,
                standardize_item_fn=standardize_scope_item,
                usage_rows=usage_rows,
                response_schema=(scope_schema if CONFIG["backend"] in {"gemini", "openai"} else None),
            )
            scope_results.extend(results)
            usage_rows.append({
                "run": f"Run {run_idx}", "pass": "pass1_scope", "batch_number": batch_num,
                "backend": CONFIG["backend"], "model": CONFIG["ollama_model"] if CONFIG["backend"] == "ollama" else (CONFIG["anthropic_model"] if CONFIG["backend"] == "anthropic" else (CONFIG["gemini_model"] if CONFIG["backend"] == "gemini" else CONFIG["openai_model"])),
                "records_in_batch": len(batch), "prompt_tokens": usage.prompt_tokens, "completion_tokens": usage.completion_tokens,
                "total_tokens": usage.total_tokens, "estimated_cost_usd": estimate_cost(usage),
            })
            save_run_checkpoint(run_idx, "scope", batch_num + 1, scope_results, pillar_results, category_results, usage_rows)
            if CONFIG["sleep_between_calls_seconds"]:
                time.sleep(CONFIG["sleep_between_calls_seconds"])

    # Repair any incomplete or dirty scope checkpoint data before proceeding
    scope_results = validate_and_repair_batch_results(
        batch=input_records,
        results=scope_results,
        pass_name="checkpoint_scope_repair",
        batch_num=0,
        run_idx=run_idx,
        system_prompt=scope_system,
        build_user_prompt_fn=build_scope_user_prompt,
        standardize_item_fn=standardize_scope_item,
        usage_rows=usage_rows,
    )
    save_run_checkpoint(run_idx, "pillar", 1 if resume_stage == "scope" else resume_batch, scope_results, pillar_results, category_results, usage_rows)

    scope_by_id = {r["record_id"]: r for r in scope_results}
    in_scope_records = [
        rec for rec in input_records
        if normalize_scope(scope_by_id[rec["record_id"]].get("raw_in_scope")) == "yes"
    ]

    pillar_batches = chunk_records(in_scope_records, CONFIG["batch_size"])
    if resume_stage in {"scope", "pillar"}:
        start_batch = 1 if resume_stage == "scope" else resume_batch
        for batch_num, batch in enumerate(tqdm(pillar_batches, desc=f"Run {run_idx} pass2_pillar"), start=1):
            if batch_num < start_batch:
                continue
            content, usage = call_model(pillar_system, build_pillar_user_prompt(batch), response_schema=(pillar_schema if CONFIG["backend"] in {"gemini", "openai"} else None))
            parsed = extract_json_object(content)
            results = [standardize_pillar_item(item) for item in parsed.get("results", [])]
            results = validate_and_repair_batch_results(
                batch=batch,
                results=results,
                pass_name="pass2_pillar",
                batch_num=batch_num,
                run_idx=run_idx,
                system_prompt=pillar_system,
                build_user_prompt_fn=build_pillar_user_prompt,
                standardize_item_fn=standardize_pillar_item,
                usage_rows=usage_rows,
                response_schema=(pillar_schema if CONFIG["backend"] in {"gemini", "openai"} else None),
            )
            pillar_results.extend(results)
            usage_rows.append({
                "run": f"Run {run_idx}", "pass": "pass2_pillar", "batch_number": batch_num,
                "backend": CONFIG["backend"], "model": CONFIG["ollama_model"] if CONFIG["backend"] == "ollama" else (CONFIG["anthropic_model"] if CONFIG["backend"] == "anthropic" else (CONFIG["gemini_model"] if CONFIG["backend"] == "gemini" else CONFIG["openai_model"])),
                "records_in_batch": len(batch), "prompt_tokens": usage.prompt_tokens, "completion_tokens": usage.completion_tokens,
                "total_tokens": usage.total_tokens, "estimated_cost_usd": estimate_cost(usage),
            })
            save_run_checkpoint(run_idx, "pillar", batch_num + 1, scope_results, pillar_results, category_results, usage_rows)
            if CONFIG["sleep_between_calls_seconds"]:
                time.sleep(CONFIG["sleep_between_calls_seconds"])

    if in_scope_records:
        pillar_results = validate_and_repair_batch_results(
            batch=in_scope_records,
            results=pillar_results,
            pass_name="checkpoint_pillar_repair",
            batch_num=0,
            run_idx=run_idx,
            system_prompt=pillar_system,
            build_user_prompt_fn=build_pillar_user_prompt,
            standardize_item_fn=standardize_pillar_item,
            usage_rows=usage_rows,
        )
    else:
        pillar_results = []

    save_run_checkpoint(run_idx, "category", 1 if resume_stage in {"scope", "pillar"} else resume_batch, scope_results, pillar_results, category_results, usage_rows)

    pillar_by_id = {r["record_id"]: r for r in pillar_results}
    category_input = []
    for rec in in_scope_records:
        p_item = pillar_by_id.get(rec["record_id"])
        pillar = normalize_predicted_pillar(p_item.get("raw_pillar")) if p_item else ""
        category_input.append({
            "record_id": rec["record_id"],
            "pillar": pillar,
            "title": rec["title"],
            "abstract": rec["abstract"],
        })

    category_batches = chunk_records(category_input, CONFIG["batch_size"])
    if resume_stage in {"scope", "pillar", "category"}:
        start_batch = 1 if resume_stage in {"scope", "pillar"} else resume_batch
        for batch_num, batch in enumerate(tqdm(category_batches, desc=f"Run {run_idx} pass3_category"), start=1):
            if batch_num < start_batch:
                continue
            print(f"Starting category batch {batch_num}...")
            content, usage = call_model(category_system, build_category_user_prompt(batch), response_schema=(category_schema if CONFIG["backend"] in {"gemini", "openai"} else None))
            print(f"Finished category batch {batch_num}")
            parsed = extract_json_object(content)
            results = [standardize_category_item(item) for item in parsed.get("results", [])]
            results = validate_and_repair_batch_results(
                batch=batch,
                results=results,
                pass_name="pass3_category",
                batch_num=batch_num,
                run_idx=run_idx,
                system_prompt=category_system,
                build_user_prompt_fn=build_category_user_prompt,
                standardize_item_fn=standardize_category_item,
                usage_rows=usage_rows,
                response_schema=(category_schema if CONFIG["backend"] in {"gemini", "openai"} else None),
            )
            category_results.extend(results)
            usage_rows.append({
                "run": f"Run {run_idx}", "pass": "pass3_category", "batch_number": batch_num,
                "backend": CONFIG["backend"], "model": CONFIG["ollama_model"] if CONFIG["backend"] == "ollama" else (CONFIG["anthropic_model"] if CONFIG["backend"] == "anthropic" else (CONFIG["gemini_model"] if CONFIG["backend"] == "gemini" else CONFIG["openai_model"])),
                "records_in_batch": len(batch), "prompt_tokens": usage.prompt_tokens, "completion_tokens": usage.completion_tokens,
                "total_tokens": usage.total_tokens, "estimated_cost_usd": estimate_cost(usage),
            })
            save_run_checkpoint(run_idx, "category", batch_num + 1, scope_results, pillar_results, category_results, usage_rows)
            if CONFIG["sleep_between_calls_seconds"]:
                time.sleep(CONFIG["sleep_between_calls_seconds"])

    if category_input:
        category_results = validate_and_repair_batch_results(
            batch=category_input,
            results=category_results,
            pass_name="checkpoint_category_repair",
            batch_num=0,
            run_idx=run_idx,
            system_prompt=category_system,
            build_user_prompt_fn=build_category_user_prompt,
            standardize_item_fn=standardize_category_item,
            usage_rows=usage_rows,
        )
    else:
        category_results = []

    category_by_id = {r["record_id"]: r for r in category_results}
    final_results = [
        normalize_three_pass_result(scope_by_id[rec["record_id"]], pillar_by_id.get(rec["record_id"]), category_by_id.get(rec["record_id"]), rules_yaml)
        for rec in input_records
    ]

    delete_run_checkpoint(run_idx)

    results_df = pd.DataFrame(final_results)
    run_df = prepared_df.copy().merge(results_df, left_on="_record_id", right_on="record_id", how="left")
    run_df = run_df.drop(columns=[c for c in ["record_id", "_record_id", "_title", "_abstract"] if c in run_df.columns])
    run_df["Pillar"] = run_df["pred_pillar"].fillna("")
    run_df["Category"] = run_df["pred_research_category"].fillna("")
    run_df["Unsure"] = run_df["pred_review_flag"].fillna("")
    run_df["Notes"] = run_df["pred_brief_reason"].fillna("")
    return run_df, pd.DataFrame(usage_rows)


def evaluate_against_solutions(run_df: pd.DataFrame, solutions_df: pd.DataFrame, run_name: str, rules_yaml: dict) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    run_eval = run_df.copy()
    sol = solutions_df.copy()
    sol["_record_id"] = sol[CONFIG["id_column"]].apply(normalize_text)
    merged = run_eval.copy()
    merged["_record_id"] = merged[CONFIG["id_column"]].apply(normalize_text)
    merged = merged.merge(sol, on="_record_id", how="left", suffixes=("", "_sol"))

    sol_pillar_series = merged["Pillar_sol"] if "Pillar_sol" in merged.columns else pd.Series([None] * len(merged))
    sol_cat_series = merged["Category_sol"] if "Category_sol" in merged.columns else pd.Series([None] * len(merged))

    merged["sol_scope_norm"] = sol_pillar_series.apply(lambda x: "yes" if normalize_text(x) else "no")
    merged["pred_scope_norm"] = merged["pred_in_scope"].apply(normalize_scope)

    pred_pillar_norm = merged["pred_pillar"].fillna("").apply(lambda x: normalize_predicted_pillar(x).strip().lower())
    sol_pillar_norm = sol_pillar_series.fillna("").apply(lambda x: normalize_solution_pillar(x).strip().lower())
    pred_cat_norm = merged["pred_research_category"].fillna("").apply(lambda x: normalize_predicted_category(x).strip().lower())
    sol_cat_norm = sol_cat_series.fillna("").apply(lambda x: normalize_solution_category(x).strip().lower())
    is_cross_cutting_solution = sol_pillar_norm.eq("cross-cutting")

    merged["scope_correct"] = merged["pred_scope_norm"] == merged["sol_scope_norm"]
    merged["pillar_correct_strict"] = (
        (merged["pred_pillar"].fillna("").astype(str).str.strip().str.lower() == sol_pillar_series.fillna("").astype(str).str.strip().str.lower())
        | is_cross_cutting_solution
        | (merged["sol_scope_norm"] != "yes")
    )
    merged["category_correct_strict"] = (
        (merged["pred_research_category"].fillna("").astype(str).str.strip().str.lower() == sol_cat_series.fillna("").astype(str).str.strip().str.lower())
        | (merged["sol_scope_norm"] != "yes")
    )
    merged["fully_correct_strict"] = merged["scope_correct"] & merged["pillar_correct_strict"] & merged["category_correct_strict"]
    merged["pillar_correct_normalized"] = ((pred_pillar_norm == sol_pillar_norm) | is_cross_cutting_solution | (merged["sol_scope_norm"] != "yes"))
    merged["category_correct_normalized"] = ((pred_cat_norm == sol_cat_norm) | (merged["sol_scope_norm"] != "yes"))
    merged["fully_correct_normalized"] = merged["scope_correct"] & merged["pillar_correct_normalized"] & merged["category_correct_normalized"]
    merged["flagged_unsure"] = merged["pred_review_flag"].fillna("").astype(str).str.lower().eq("unsure")
    merged["flagged_borderline"] = merged["pred_review_flag"].fillna("").astype(str).str.lower().eq("borderline")

    summary = {
        "run": run_name,
        "benchmark_version": rules_yaml.get("benchmark_version", "unknown"),
        "backend": CONFIG["backend"],
        "model": CONFIG["ollama_model"] if CONFIG["backend"] == "ollama" else (CONFIG["anthropic_model"] if CONFIG["backend"] == "anthropic" else (CONFIG["gemini_model"] if CONFIG["backend"] == "gemini" else CONFIG["openai_model"])),
        "benchmark_mode": "three_pass",
        "batch_size": CONFIG["batch_size"],
        "temperature": CONFIG["temperature"],
        "n_records": int(len(merged)),
        "scope_accuracy": float(merged["scope_correct"].mean()),
        "full_accuracy_strict": float(merged["fully_correct_strict"].mean()),
        "pillar_accuracy_on_in_scope_strict": float(merged.loc[merged["sol_scope_norm"] == "yes", "pillar_correct_strict"].mean()) if (merged["sol_scope_norm"] == "yes").any() else None,
        "category_accuracy_on_in_scope_strict": float(merged.loc[merged["sol_scope_norm"] == "yes", "category_correct_strict"].mean()) if (merged["sol_scope_norm"] == "yes").any() else None,
        "full_accuracy_normalized": float(merged["fully_correct_normalized"].mean()),
        "pillar_accuracy_on_in_scope_normalized": float(merged.loc[merged["sol_scope_norm"] == "yes", "pillar_correct_normalized"].mean()) if (merged["sol_scope_norm"] == "yes").any() else None,
        "category_accuracy_on_in_scope_normalized": float(merged.loc[merged["sol_scope_norm"] == "yes", "category_correct_normalized"].mean()) if (merged["sol_scope_norm"] == "yes").any() else None,
        "unsure_rate": float(merged["flagged_unsure"].mean()),
        "borderline_rate": float(merged["flagged_borderline"].mean()),
    }
    return merged, summary


def build_error_summary_counts(item_level_analysis: pd.DataFrame) -> Dict[str, int]:
    if item_level_analysis.empty:
        return {}
    df = item_level_analysis.copy()
    true_scope = df["Pillar_sol"].apply(lambda x: "yes" if normalize_text(x) else "no") if "Pillar_sol" in df.columns else pd.Series(["no"] * len(df))
    pred_scope = df["pred_in_scope"].apply(normalize_scope) if "pred_in_scope" in df.columns else pd.Series([""] * len(df))
    pillar_ok = df["pillar_correct_normalized"] if "pillar_correct_normalized" in df.columns else pd.Series([False] * len(df))
    category_ok = df["category_correct_normalized"] if "category_correct_normalized" in df.columns else pd.Series([False] * len(df))
    fully_correct = df["fully_correct_normalized"] if "fully_correct_normalized" in df.columns else pd.Series([False] * len(df))
    unsure = df["pred_review_flag"].fillna("").astype(str).str.lower().eq("unsure") if "pred_review_flag" in df.columns else pd.Series([False] * len(df))
    borderline = df["pred_review_flag"].fillna("").astype(str).str.lower().eq("borderline") if "pred_review_flag" in df.columns else pd.Series([False] * len(df))
    pred_pillar = df.get("pred_pillar", pd.Series([""] * len(df))).fillna("").astype(str)
    pred_cat = df.get("pred_research_category", pd.Series([""] * len(df))).fillna("").astype(str)
    return {
        "total_rows": int(len(df)),
        "true_in_scope": int(true_scope.eq("yes").sum()),
        "true_out_of_scope": int(true_scope.eq("no").sum()),
        "wrong_scope": int(pred_scope.ne(true_scope).sum()),
        "right_scope": int(pred_scope.eq(true_scope).sum()),
        "right_scope_wrong_pillar": int((pred_scope.eq(true_scope) & true_scope.eq("yes") & (~pillar_ok)).sum()),
        "right_scope_and_pillar_wrong_category": int((pred_scope.eq(true_scope) & true_scope.eq("yes") & pillar_ok & (~category_ok)).sum()),
        "fully_correct_normalized_count": int(fully_correct.sum()),
        "unsure_flagged_count": int(unsure.sum()),
        "borderline_flagged_count": int(borderline.sum()),
        "unsure_and_fully_correct_count": int((unsure & fully_correct).sum()),
        "unsure_and_wrong_count": int((unsure & (~fully_correct)).sum()),
        "missing_predicted_pillar_on_in_scope": int((true_scope.eq("yes") & pred_pillar.eq("")).sum()),
        "missing_predicted_category_on_in_scope": int((true_scope.eq("yes") & pred_cat.eq("")).sum()),
        "validation_status_normalized": int(df["validation_status"].eq("normalized").sum()) if "validation_status" in df.columns else 0,
        "validation_status_valid": int(df["validation_status"].eq("valid").sum()) if "validation_status" in df.columns else 0,
    }


def build_item_level_export(df: pd.DataFrame, run_name: str) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    out = pd.DataFrame(index=df.index.copy())
    out["run"] = run_name

    for col in ["Publication ID", "DOI", "PMID", "PMCID", "ISBN", "Title", "Abstract"]:
        out[col] = df[col] if col in df.columns else ""

    out["pred_in_scope"] = df["pred_in_scope"] if "pred_in_scope" in df.columns else ""
    out["pred_pillar"] = df["pred_pillar"] if "pred_pillar" in df.columns else ""
    out["pred_research_category"] = df["pred_research_category"] if "pred_research_category" in df.columns else ""
    out["pred_review_flag"] = df["pred_review_flag"] if "pred_review_flag" in df.columns else ""
    out["pred_brief_reason"] = df["pred_brief_reason"] if "pred_brief_reason" in df.columns else ""
    out["Pillar_sol"] = df["Pillar_sol"] if "Pillar_sol" in df.columns else ""
    out["Category_sol"] = df["Category_sol"] if "Category_sol" in df.columns else ""
    out["sol_scope_norm"] = df["sol_scope_norm"] if "sol_scope_norm" in df.columns else ""
    out["scope_correct"] = df["scope_correct"] if "scope_correct" in df.columns else ""
    out["pillar_correct_normalized"] = df["pillar_correct_normalized"] if "pillar_correct_normalized" in df.columns else ""
    out["category_correct_normalized"] = df["category_correct_normalized"] if "category_correct_normalized" in df.columns else ""
    out["fully_correct_normalized"] = df["fully_correct_normalized"] if "fully_correct_normalized" in df.columns else ""
    out["validation_status"] = df["validation_status"] if "validation_status" in df.columns else ""
    out["validation_notes"] = df["validation_notes"] if "validation_notes" in df.columns else ""

    return out.reset_index(drop=True)


def order_summary_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    ordered = [c for c in SUMMARY_VISIBLE_COLUMNS if c in df.columns]
    ordered += [c for c in df.columns if c not in ordered]
    return df[ordered]


def build_run_presentation_sheet(analysis_df: pd.DataFrame) -> pd.DataFrame:
    df = analysis_df.copy()
    out = pd.DataFrame()
    for col in ["Publication ID", "DOI", "PMID", "PMCID", "ISBN", "Title", "Abstract"]:
        out[col] = df[col] if col in df.columns else ""
    out["Scope"] = df["pred_in_scope"] if "pred_in_scope" in df.columns else ""
    out["Pillar"] = df["Pillar"] if "Pillar" in df.columns else df.get("pred_pillar", "")
    out["Category"] = df["Category"] if "Category" in df.columns else df.get("pred_research_category", "")
    out["Unsure"] = df["Unsure"] if "Unsure" in df.columns else df.get("pred_review_flag", "")
    out["Notes"] = df["Notes"] if "Notes" in df.columns else df.get("pred_brief_reason", "")
    out["solution_pillar_normalized"] = df["Pillar_sol"].apply(normalize_solution_pillar) if "Pillar_sol" in df.columns else ""
    out["true_scope_for_display"] = out["solution_pillar_normalized"].apply(lambda x: "yes" if normalize_text(x) else "no")
    out["solution_category_normalized"] = df["Category_sol"].apply(normalize_solution_category) if "Category_sol" in df.columns else ""
    return out


def apply_output_formatting_to_sheet(ws) -> None:
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    scope_col = headers.get("Scope")
    pillar_col = headers.get("Pillar")
    category_col = headers.get("Category")
    unsure_col = headers.get("Unsure")
    sol_pillar_col = headers.get("solution_pillar_normalized")
    true_scope_col = headers.get("true_scope_for_display")
    sol_category_col = headers.get("solution_category_normalized")

    if CONFIG.get("hide_technical_columns"):
        for header, col_idx in headers.items():
            if header not in MAIN_DISPLAY_COLUMNS:
                ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    for row_idx in range(2, ws.max_row + 1):
        unsure_val = normalize_text(ws.cell(row=row_idx, column=unsure_col).value).lower() if unsure_col else ""
        if scope_col and true_scope_col:
            pred_scope = normalize_scope(ws.cell(row=row_idx, column=scope_col).value)
            true_scope = normalize_scope(ws.cell(row=row_idx, column=true_scope_col).value)
            ws.cell(row=row_idx, column=scope_col).fill = GREEN_FILL if pred_scope == true_scope else RED_FILL
        if pillar_col and sol_pillar_col:
            pred_val = normalize_predicted_pillar(ws.cell(row=row_idx, column=pillar_col).value).strip().lower()
            sol_val = normalize_solution_pillar(ws.cell(row=row_idx, column=sol_pillar_col).value).strip().lower()
            ws.cell(row=row_idx, column=pillar_col).fill = GREEN_FILL if pred_val == sol_val else RED_FILL
        if category_col and sol_category_col:
            pred_val = normalize_predicted_category(ws.cell(row=row_idx, column=category_col).value).strip().lower()
            sol_val = normalize_solution_category(ws.cell(row=row_idx, column=sol_category_col).value).strip().lower()
            ws.cell(row=row_idx, column=category_col).fill = GREEN_FILL if pred_val == sol_val else RED_FILL
        if unsure_col and unsure_val in {"unsure", "borderline"}:
            ws.cell(row=row_idx, column=unsure_col).fill = YELLOW_FILL


def safe_load_workbook(path: str):
    if not Path(path).exists():
        wb = Workbook()
        wb.active.title = "Summary"
        wb.save(path)
    try:
        wb = load_workbook(path)
    except (BadZipFile, KeyError, OSError, FileNotFoundError):
        wb = Workbook()
        wb.active.title = "Summary"
        wb.save(path)
        wb = load_workbook(path)
    return wb


def ensure_visible_sheets(wb) -> None:
    if not wb.sheetnames:
        wb.create_sheet("Summary")
    for name in wb.sheetnames:
        wb[name].sheet_state = "visible"
    if "Summary" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Summary")
    else:
        wb.active = 0


def replace_sheet_with_df(wb, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    return ws


def read_sheet_df(path: str, sheet_name: str) -> pd.DataFrame:
    if not Path(path).exists():
        return pd.DataFrame()
    try:
        xls = pd.ExcelFile(path)
        if sheet_name in xls.sheet_names:
            return pd.read_excel(path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()
    return pd.DataFrame()


def append_or_create_sheet(wb, path: str, sheet_name: str, new_df: pd.DataFrame, order_fn=None):
    existing_df = read_sheet_df(path, sheet_name)
    combined_df = pd.concat([existing_df, new_df], ignore_index=True, sort=False) if not existing_df.empty else new_df.copy()
    if order_fn is not None:
        combined_df = order_fn(combined_df)
    return replace_sheet_with_df(wb, sheet_name, combined_df)


def write_output_excel(
    test_df: pd.DataFrame,
    solutions_df: pd.DataFrame,
    run_tabs: Dict[str, pd.DataFrame],
    summaries: List[Dict[str, Any]],
    item_level_analysis: pd.DataFrame,
) -> None:
    path = CONFIG["output_file"]
    wb = safe_load_workbook(path)

    replace_sheet_with_df(wb, "Test data", test_df)
    replace_sheet_with_df(wb, "Solutions", solutions_df)

    for sheet_name, analysis_df in run_tabs.items():
        ws_run = replace_sheet_with_df(wb, sheet_name[:31], build_run_presentation_sheet(analysis_df))
        apply_output_formatting_to_sheet(ws_run)

    append_or_create_sheet(wb, path, "Summary", order_summary_columns(pd.DataFrame(summaries)))
    append_or_create_sheet(wb, path, "Item-level analysis", item_level_analysis)

    desired = (
        ["Test data", "Solutions"]
        + sorted([s for s in wb.sheetnames if re.fullmatch(r"Run\s+\d+", s)], key=lambda x: int(x.split()[1]))
        + ["Summary", "Item-level analysis"]
    )
    wb._sheets = [wb[s] for s in desired if s in wb.sheetnames]

    if CONFIG.get("hide_technical_columns") and "Summary" in wb.sheetnames:
        ws_summary = wb["Summary"]
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws_summary[1])}
        for header, col_idx in headers.items():
            ws_summary.column_dimensions[get_column_letter(col_idx)].hidden = header not in SUMMARY_VISIBLE_COLUMNS

    ensure_visible_sheets(wb)
    wb.save(path)


def main() -> None:
    load_env()
    test_df, solutions_df, rules_text, rules_yaml = load_input_data()
    prepared_df = prepare_records(test_df)

    run_tabs: Dict[str, pd.DataFrame] = {}
    all_summaries: List[Dict[str, Any]] = []
    all_item_level: List[pd.DataFrame] = []

    resume_run_number = get_resume_run_number()
    if resume_run_number is not None:
        print(f"Resuming incomplete run {resume_run_number} from checkpoint...")
        start_run_number = resume_run_number
    else:
        existing_run_sheets = []
        if Path(CONFIG["output_file"]).exists():
            try:
                wb = load_workbook(CONFIG["output_file"], read_only=True)
                existing_run_sheets = [s for s in wb.sheetnames if re.fullmatch(r"Run\s+\d+", s)]
            except Exception:
                existing_run_sheets = []
        start_run_number = (max([int(s.split()[1]) for s in existing_run_sheets]) + 1) if existing_run_sheets else 1

    for offset in range(int(CONFIG["repetitions"])):
        run_idx = start_run_number + offset
        run_name = f"Run {run_idx}"
        run_df, usage_df = run_one_experiment(prepared_df, rules_text, rules_yaml, run_idx)
        analysis_df, summary = evaluate_against_solutions(run_df, solutions_df, run_name, rules_yaml)
        run_tabs[run_name] = analysis_df
        all_summaries.append({**summary, **build_error_summary_counts(analysis_df)})
        all_item_level.append(build_item_level_export(analysis_df, run_name))

    combined_item_level = pd.concat(all_item_level, ignore_index=True) if all_item_level else pd.DataFrame()

    write_output_excel(
        test_df=test_df,
        solutions_df=solutions_df,
        run_tabs=run_tabs,
        summaries=all_summaries,
        item_level_analysis=combined_item_level,
    )
    print(f"Done. Wrote results to: {CONFIG['output_file']}")


if __name__ == "__main__":
    main()
